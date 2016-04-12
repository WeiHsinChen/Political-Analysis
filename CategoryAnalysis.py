import nltk
from nltk import word_tokenize
from nltk.stem.snowball import SnowballStemmer
from nltk.stem import WordNetLemmatizer 
from openpyxl import Workbook
from openpyxl import load_workbook

class CategoryAnalysis:
  def __init__(self):
    """
    Construct a new 'CategoryAnalysis' object.
    """
    self.domain_dict = None
  def parse_raw_dict(self):
    """ 
    Parse all words in Academic Vocabulary Lists and extract keywords for each domain.
    """
    fName = "datasets/allWords.xlsx"
    wb = load_workbook(filename = fName)
    ws = wb.get_sheet_by_name("list")
    wb_res = Workbook()
    ws_res = wb_res.active
    record = {}
    stemmer = SnowballStemmer("english")
    wordnet_lemmatizer = WordNetLemmatizer()
    i = 2
    j = 1
    while True:
      try:
        w = str(ws['D'+str(i)].value)
        domains = str(ws['k'+str(i)].value).replace('+', ',')
        w_stem = stemmer.stem(w.lower())
        # w_stem = wordnet_lemmatizer.lemmatize(w.lower())
        if w == 'None':
          break
        if domains == 'None':
          i += 1
          continue

        # print i, w, w_stem, domains
        if w_stem not in record:
          record[w_stem] = [j, set(domains.split(','))]
          # print i, 'w_stem: ', w_stem, record[w_stem][1]
          ws_res['A'+str(j)] = w_stem
          ws_res['B'+str(j)] = domains
          j += 1
        else:
          domains_array = domains.split(',')
          for d in domains_array:
            if d not in record[w_stem][1]:
              record[w_stem][1].add(d)
              ws_res['B'+str(record[w_stem][0])] = str(ws_res['B'+str(record[w_stem][0])].value) + ','+ d
        i += 1
      except:
        i += 1
    wb_res.save('datasets/domain_dict.xlsx')

  def get_domain_dict(self):
    """ 
    Parse keywords of each domain from xlsx files.
    """
    res = {}
    fName = "datasets/domain_dict.xlsx"
    wb = load_workbook(filename = fName)
    ws = wb.get_sheet_by_name("Sheet")

    i = 1
    while True:
      w = str(ws['A'+str(i)].value)
      domains = str(ws['B'+str(i)].value).split(',')
      if w == 'None':
        break
      res[w] = domains
      i += 1

    self.domain_dict = res

  def predict_domain(self, text):
    """
    Predict domain of a new tweet.

    :param text: The text of the tweet
    :return: The domain of the tweet, categorize text into 'His', 'Soc', 'Sci', 'Edu', 'Rel', 'Med', 'Law', 'Bus' or 'Neu' based on generative model 
    """
    if not self.domain_dict:
      self.get_domain_dict()

    count_dict = {}
    text_array = nltk.word_tokenize(text.strip())
    stemmer = SnowballStemmer("english")
    # wordnet_lemmatizer = WordNetLemmatizer()

    for w in text_array:
      w_stem = stemmer.stem(w.lower())
      # w_stem = wordnet_lemmatizer.lemmatize(w.lower())
      if w_stem in self.domain_dict:
        for d in self.domain_dict[w_stem]:
          if d not in count_dict:
            count_dict[d] = 0
          count_dict[d] += 1

    # Does not belong to any domain
    if len(count_dict) == 0:
      return ['Neu']
    else: 
      max_freq = 0
      max_domains = ''
      for d, f in count_dict.items():
        if f > max_freq:
          max_freq = f
          max_domains = [d]
        elif f == max_freq:
          max_domains.append(d)

      return max_domains


# tmp = CategoryAnalysis()
# # tmp.parse_raw_dict()
# print tmp.predict_domain("trump's social policy is stupid.")
